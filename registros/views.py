from rest_framework.permissions import IsAuthenticated, IsAdminUser
from backend.permissions import IsAdminUserOrReadOnly

# Para hacer transacciones atómicas. Asegura que todas las operaciones se completen
from django.db import transaction

from rest_framework import generics, views
from rest_framework.parsers import FileUploadParser
from rest_framework.response import Response
from rest_framework.decorators import api_view, permission_classes 

# Procesar en chunks más pequeños pero con procesamiento paralelo
from concurrent.futures import ThreadPoolExecutor

from .serializers import IngresoSerializer, EgresoSerializer, TitulacionSerializer, LiberacionInglesSerializer
from .models import Ingreso, Egreso, Titulacion, LiberacionIngles
from .periodos import getPeriodoActual

from personal.models import Personal, obtenerFechaNac, obtenerGenero
from alumnos.models import Alumno
from carreras.models import Carrera
from planes.models import Plan

import openpyxl
import re
import pandas as pd
import numpy as np
import multiprocessing
import logging
import time
import gc
import psutil

logger = logging.getLogger(__name__)

### INGRESO
class IngresoList(generics.ListCreateAPIView):
    queryset = Ingreso.objects.all()
    serializer_class = IngresoSerializer
    permission_classes = [IsAuthenticated&IsAdminUserOrReadOnly]

class IngresoDetail(generics.RetrieveUpdateDestroyAPIView):
    queryset = Ingreso.objects.all()
    serializer_class = IngresoSerializer
    permission_classes = [IsAuthenticated&IsAdminUserOrReadOnly]

# FORMATO DE EXCEL [CURP, NO_CONTROL, PATERNO, MATERNO, NOMBRE, CARRERA, (PERIODO+TIPO)]
class IngresoUpload(views.APIView):
    parser_classes = [FileUploadParser]
    permission_classes = [IsAuthenticated & IsAdminUser]
    CHUNK_SIZE = 300  # Reducido para archivos pequeños/medianos | El más adecuado para la aplicación

    def validate_data(self, df):
        """Valida el DataFrame y extrae el periodo de la última columna"""
        if df.empty:
            raise Exception('Archivo vacío')

        # Validar número de columnas
        if len(df.columns) != 7:
            raise Exception('Número incorrecto de columnas')

        # Obtener y validar periodo de la última columna
        periodo_col = str(df.columns[-1])
        if not re.match(r'^[12][0-9]{3}[13]$', periodo_col):
            raise Exception(f'Formato de periodo inválido: {periodo_col}')

        # Renombrar columnas
        df.columns = ['curp', 'no_control', 'paterno', 'materno', 'nombre', 'carrera', 'tipo']

        # Añadir columna de periodo
        df['periodo'] = periodo_col

        # Limpiar datos
        df = df.apply(lambda x: x.str.strip() if x.dtype == "object" else x)

        # Validar datos requeridos
        required_fields = ['curp', 'no_control', 'nombre', 'carrera', 'tipo']
        df = df.dropna(subset=required_fields)

        return df

    def get_cached_data(self):
        """Obtiene y cachea los datos necesarios de manera más eficiente"""
        with transaction.atomic():
            # Obtener ingresos existentes usando iterator() para mejor memoria
            existing_ingresos = set(
                Ingreso.objects.values_list(
                    'alumno_id', 'periodo', 'tipo'
                ).iterator()
            )

            # Obtener alumnos con datos relacionados en una sola consulta
            existing_alumnos = {
                a.no_control: a 
                for a in Alumno.objects.select_related(
                    'plan',
                    'plan__carrera'
                ).only(
                    'no_control',
                    'plan__carrera__clave',
                    'curp_id'
                ).all()
            }

            # Obtener solo los datos necesarios de carreras
            carreras = {
                c.clave: c 
                for c in Carrera.objects.only('clave').all()
            }

            # Obtener planes con su carrera relacionada
            planes = {
                p.carrera.clave: p 
                for p in Plan.objects.select_related('carrera').only(
                    'clave',
                    'carrera'
                ).all()
            }

            return existing_ingresos, existing_alumnos, carreras, planes

    def get_optimal_workers(self, total_records):
        """Determinar número óptimo de workers según tamaño de datos"""
        if total_records < 500:
            return 2
        elif total_records < 2000:
            return min(4, multiprocessing.cpu_count())
        else:
            return min(8, multiprocessing.cpu_count())

    def get_optimal_chunk_size(self, total_records):
        """Determinar tamaño óptimo de chunk según registros totales"""
        if total_records < 500:
            return 100
        elif total_records < 2000:
            return 300
        else:
            return 500

    def bulk_create_with_progress(self, model, objects, batch_size=100):
        """Crear objetos en lotes con mejor manejo de memoria"""
        if not objects:
            return 0
        
        total = len(objects)
        created = 0
        
        for i in range(0, total, batch_size):
            batch = objects[i:i + batch_size]
            created += len(
                model.objects.bulk_create(
                    batch,
                    ignore_conflicts=True
                )
            )
            del batch  # Liberar memoria
            gc.collect()  # Forzar recolección de basura
        
        return created

    def get_optimal_chunk_configuration(self, df):
        """Configuración optimizada de chunks basada en recursos del sistema"""
        try:
            # Calcular memoria disponible y tamaño de registros
            memory = psutil.virtual_memory()
            available_memory = memory.available
            record_size = df.memory_usage(deep=True).sum() / len(df)
            total_records = len(df)

            # Calcular tamaño óptimo de chunk basado en memoria disponible
            # Usar solo 30% de la memoria disponible para ser conservadores
            memory_based_size = int((available_memory * 0.3) / record_size)

            # Aplicar límites basados en número de registros
            if total_records < 500:
                base_size = 100
            elif total_records < 2000:
                base_size = 300
            else:
                base_size = 500

            # Tomar el mínimo entre el tamaño basado en memoria y el basado en registros
            optimal_size = min(memory_based_size, base_size)

            logger.info(
                f"Configuración de chunks: "
                f"Registros totales: {total_records}, "
                f"Memoria disponible: {available_memory / (1024*1024):.2f}MB, "
                f"Tamaño por registro: {record_size / 1024:.2f}KB, "
                f"Tamaño chunk: {optimal_size}"
            )

            return optimal_size

        except Exception as ex:
            logger.warning(f"Error al calcular tamaño de chunk: {str(ex)}")
            # Fallback a configuración estática si hay error
            return self.get_optimal_chunk_size(len(df))

    def post(self, request, filename, format=None):
        start_time = time.time()
        file_obj = request.data['file']
        results = {"errors": [], "created": 0}

        try:
            # Tu código existente de lectura
            df = pd.read_excel(
                file_obj,
                header=0,
                engine='openpyxl',
                na_filter=False,
                usecols="A:G",
                converters={
                    'curp': lambda x: str(x).strip() if pd.notna(x) else '',
                    'no_control': lambda x: str(x).strip() if pd.notna(x) else '',
                    'paterno': lambda x: str(x).strip() if pd.notna(x) else '',
                    'materno': lambda x: str(x).strip() if pd.notna(x) else '',
                    'nombre': lambda x: str(x).strip() if pd.notna(x) else '',
                    'carrera': lambda x: str(x).strip() if pd.notna(x) else ''
                }
            )

            df = self.validate_data(df)
            existing_data = self.get_cached_data()

            # Usar la nueva configuración optimizada
            chunk_size = self.get_optimal_chunk_configuration(df)
            num_workers = self.get_optimal_workers(len(df))
            
            chunks = [
                df.iloc[i:i + chunk_size] 
                for i in range(0, len(df), chunk_size)
            ]
            
            all_personal = []
            all_alumnos = []
            all_ingresos = []

            # Procesamiento paralelo optimizado
            with ThreadPoolExecutor(max_workers=num_workers) as executor:
                futures = [
                    executor.submit(self.process_chunk, chunk, existing_data, results)
                    for chunk in chunks
                ]

                for future in futures:
                    personal_chunk, alumnos_chunk, ingresos_chunk = future.result()
                    all_personal.extend(personal_chunk)
                    all_alumnos.extend(alumnos_chunk)
                    all_ingresos.extend(ingresos_chunk)

            # Tu código existente de bulk_create
            try:
                with transaction.atomic():
                    results['created'] += self.bulk_create_with_progress(Personal, all_personal)
                    results['created'] += self.bulk_create_with_progress(Alumno, all_alumnos)
                    results['created'] += self.bulk_create_with_progress(Ingreso, all_ingresos)

            except Exception as ex:
                results['errors'].append({
                    'type': str(type(ex)),
                    'message': 'Error en bulk create: ' + str(ex)
                })

            # Logging del rendimiento
            processing_time = time.time() - start_time
            logger.info(
                f"Archivo procesado: {filename} "
                f"Tiempo: {processing_time:.2f}s "
                f"Registros: {len(df)} "
                f"Creados: {results['created']}"
            )

            # Limpiar memoria final
            del df
            gc.collect()

        except Exception as ex:
            results['errors'].append({
                'type': str(type(ex)),
                'message': 'Error en el procesamiento: ' + str(ex)
            })

        return Response(status=200, data=results)

    def process_chunk(self, chunk_data, existing_data, results):
        """Procesa un chunk de datos y retorna las listas de objetos a crear"""
        personal_chunk = []
        alumnos_chunk = []
        ingresos_chunk = []

        existing_ingresos, existing_alumnos, carreras, planes = existing_data

        try:
            for index, row in chunk_data.iterrows():
                try:
                    # Las mismas validaciones que ya tienes
                    Personal.validate_curp(row['curp'])
                    Alumno.validate_nocontrol(row['no_control'])

                    if row['carrera'] not in carreras:
                        results['errors'].append({
                            'type': 'CarreraDoesNotExist',
                            'message': f'La carrera {row["carrera"]} no existe',
                            'row_index': index + 2
                        })
                        continue

                    # Tu lógica existente de verificación de alumno
                    alumno = existing_alumnos.get(row['no_control'])
                    if alumno:
                        if alumno.plan.carrera.clave != row['carrera']:
                            results['errors'].append({
                                'type': 'Carrera',
                                'message': "Carrera no coincide",
                                'row_index': index + 2
                            })
                            continue
                    else:
                        personal_chunk.append(
                            Personal(
                                curp=row['curp'],
                                paterno=row['paterno'],
                                materno=row['materno'],
                                nombre=row['nombre'],
                                fecha_nacimiento=obtenerFechaNac(row['curp']),
                                genero=obtenerGenero(row['curp'])
                            )
                        )

                        plan = planes.get(row['carrera'])
                        alumnos_chunk.append(
                            Alumno(
                                no_control=row['no_control'],
                                curp_id=row['curp'],
                                plan=plan
                            )
                        )

                    # Tu lógica existente de creación de ingreso
                    ingreso_key = (row['no_control'], row['periodo'], row['tipo'])
                    if ingreso_key not in existing_ingresos:
                        ingreso = Ingreso(
                            alumno_id=row['no_control'],
                            periodo=row['periodo'],
                            tipo=row['tipo']
                        )
                        ingreso.calcular_num_semestre()
                        ingreso.full_clean()
                        ingresos_chunk.append(ingreso)

                except Exception as ex:
                    results['errors'].append({
                        'type': str(type(ex)),
                        'message': str(ex),
                        'row_index': index + 2
                    })

            # Liberar memoria del chunk procesado
            del chunk_data
            gc.collect()

        except Exception as ex:
            logger.error(f"Error en procesamiento de chunk: {str(ex)}")
            raise

        return personal_chunk, alumnos_chunk, ingresos_chunk

### EGRESO
class EgresoList(generics.ListCreateAPIView):
    queryset = Egreso.objects.all()
    serializer_class = EgresoSerializer
    permission_classes = [IsAuthenticated&IsAdminUserOrReadOnly]

class EgresoDetail(generics.RetrieveUpdateDestroyAPIView):
    queryset = Egreso.objects.all()
    serializer_class = EgresoSerializer
    permission_classes = [IsAuthenticated&IsAdminUserOrReadOnly]

# FORMATO DE EXCEL [NO_CONTROL, PERIODO]
class EgresoUpload(views.APIView):
    parser_classes = [FileUploadParser]
    permission_classes = [IsAuthenticated&IsAdminUser]

    def to_dict(self, row):
        # regresa None si el renglon son solo celdas vacias
        for cell in row:
            if cell.value is not None:
                break
            return None

        if row[0].value is None:
            raise Exception('Se necesita un no. de control')

        data = {
            'no_control': str(row[0].value).strip(),
        }
        return data

    def post(self, request, filename, format=None):
        ESTRUCTURA = [(r'^no_control$', 'NO_CONTROL'), (r'^[12][0-9]{3}[13]$', 'PERIODO')]
        # Obtiene el archivo enviado en la solicitud HTTP
        file_obj = request.data['file']

        # Carga el archivo Excel en un objeto Workbook de openpyxl, 
        # con data_only=True para obtener los valores calculados en lugar de las fórmulas
        wb = openpyxl.load_workbook(file_obj, data_only=True)

        # Selecciona la hoja activa del archivo Excel
        ws = wb.active

        results = {"errors": [], "created": 0}
        header_row = ws['A1':'B1'][0] # ws['A1':'B1'] regresa una tupla de renglones, pero solo necesitamos la primera

        # VALIDAR ESTRUCTURA DEL ARCHIVO COMO:
        # no_control | periodo
        for i, expresion in enumerate(ESTRUCTURA):
            match = re.match(expresion[0], str(header_row[i].value).lower())
            if match is None:
                return Response(status=400, data={'message': f'Se esperaba el campo {expresion[i]} pero se obtuvo {header_row[i].value}'})

        for row in ws.iter_rows(min_row=2):
            try:
                data = self.to_dict(row)
                if data is None:
                    continue
                # Buscar el alumno con el número de control proporcionado en los datos
                alumno = Alumno.objects.get(pk=(data['no_control']))
                # Se crea un registro de egreso con el alumno y el periodo proporcionados
                egresado = Egreso.objects.create(periodo=str(header_row[1].value), alumno=alumno)
                egresado.save()
                results['created'] += 1
            except Alumno.DoesNotExist as ex:
                results['errors'].append({'type': str(type(ex)), 'message': f'No se encontro un alumno con no. de control {data["no_control"]}', 'row_index': row[0].row})
            except Exception as ex:
                results['errors'].append({'type': str(type(ex)), 'message': str(ex), 'row_index': row[0].row})
        return Response(status=200, data=results)

### TITULACION
class TitulacionList(generics.ListCreateAPIView):
    queryset = Titulacion.objects.all()
    serializer_class = TitulacionSerializer
    permission_classes = [IsAuthenticated&IsAdminUserOrReadOnly]

class TitulacionDetail(generics.RetrieveUpdateDestroyAPIView):
    queryset = Titulacion.objects.all()
    serializer_class = TitulacionSerializer
    permission_classes = [IsAuthenticated&IsAdminUserOrReadOnly]

# FORMATO DE EXCEL [NO_CONTROL, (PERIODO+TIPO)]
class TitulacionUpload(views.APIView):
    parser_classes = [FileUploadParser]
    permission_classes = [IsAuthenticated&IsAdminUser]

    def to_dict(self, row):
        # regresa None si el renglon son solo celdas vacias
        for cell in row:
            if cell.value is not None:
                break
            return None

        if row[0].value is None:
            raise Exception('Se necesita un no. de control')
        if row[1].value is None:
            raise Exception('Se necesita el tipo de titulación')

        data = {
            'no_control': str(row[0].value).strip(),
            'tipo_titulacion': str(row[1].value).strip()[0:2],
        }
        return data

    def post(self, request, filename, format=None):
        ESTRUCTURA = [(r'^no_control$', 'NO_CONTROL'), (r'^[12][0-9]{3}[13]$', 'NUMERO DE PERIODO')]
        # Obtiene el archivo enviado en la solicitud HTTP
        file_obj = request.data['file']

        # Carga el archivo Excel en un objeto Workbook de openpyxl, 
        # con data_only=True para obtener los valores calculados en lugar de las fórmulas
        wb = openpyxl.load_workbook(file_obj, data_only=True)

        # Selecciona la hoja activa del archivo Excel
        ws = wb.active

        results = {"errors": [], "created": 0}
        header_row = ws['A1':'B1'][0] # ws['A1':'B1'] regresa una tupla de renglones, pero solo necesitamos la primera

        # VALIDAR ESTRUCTURA DEL ARCHIVO COMO:
        # no_control | periodo
        for i, expresion in enumerate(ESTRUCTURA):
            match = re.match(expresion[0], str(header_row[i].value).lower())
            if match is None:
                return Response(status=400, data={'message': f'Se esperaba el campo {expresion[i]} pero se obtuvo {header_row[i].value}'})

        for row in ws.iter_rows(min_row=2):
            try:
                data = self.to_dict(row)
                if data is None:
                    continue
                alumno = Alumno.objects.get(pk=data['no_control'])
                titulacion, created = Titulacion.objects.get_or_create(periodo=str(header_row[1].value), tipo=data['tipo_titulacion'], alumno=alumno)
                if created:
                    results['created'] += 1
            except Alumno.DoesNotExist as ex:
                results['errors'].append({'type': str(type(ex)), 'message': f'No se encontro un alumno con no. de control {data["no_control"]}', 'row_index': row[0].row})
            except Exception as ex:
                results['errors'].append({'type': str(type(ex)), 'message': str(ex), 'row_index': row[0].row})
        return Response(status=200, data=results)

### LIBERACION DE INGLES
class LiberacionInglesList(generics.ListCreateAPIView):
    queryset = LiberacionIngles.objects.all()
    serializer_class = LiberacionInglesSerializer
    permission_classes = [IsAuthenticated&IsAdminUserOrReadOnly]

class LiberacionInglesDetail(generics.RetrieveUpdateDestroyAPIView):
    queryset = LiberacionIngles.objects.all()
    serializer_class = LiberacionInglesSerializer
    permission_classes = [IsAuthenticated&IsAdminUserOrReadOnly]

# FORMATO DE EXCEL [NO_CONTROL, PERIODO]
class LiberacionInglesUpload(views.APIView):
    parser_classes = [FileUploadParser]
    permission_classes = [IsAuthenticated&IsAdminUser]

    def to_dict(self, row):
        # regresa None si el renglon son solo celdas vacias
        for cell in row:
            if cell.value is not None:
                break
            return None

        if row[0].value is None:
            raise Exception('Se necesita un no. de control')

        data = {
            'no_control': str(row[0].value).strip(),
        }
        return data

    def post(self, request, filename, format=None):
        ESTRUCTURA = [(r'^no_control$', 'NO_CONTROL'), (r'^[12][0-9]{3}[13]$', 'NUMERO DE PERIODO')]
        # Obtiene el archivo enviado en la solicitud HTTP
        file_obj = request.data['file']

        # Carga el archivo Excel en un objeto Workbook de openpyxl, 
        # con data_only=True para obtener los valores calculados en lugar de las fórmulas
        wb = openpyxl.load_workbook(file_obj, data_only=True)

        # Selecciona la hoja activa del archivo Excel
        ws = wb.active

        results = {"errors": [], "created": 0}
        header_row = ws['A1':'B1'][0] # ws['A1':'B1'] regresa una tupla de renglones, pero solo necesitamos la primera

        # VALIDAR ESTRUCTURA DEL ARCHIVO COMO:
        # no_control | periodo
        for i, expresion in enumerate(ESTRUCTURA):
            match = re.match(expresion[0], str(header_row[i].value).lower())
            if match is None:
                return Response(status=400, data={'message': f'Se esperaba el campo {expresion[i]} pero se obtuvo {header_row[i].value}'})

        for row in ws.iter_rows(min_row=2):
            try:
                data = self.to_dict(row)
                if data is None:
                    continue
                alumno = Alumno.objects.get(pk=data['no_control'])
                liberacion, created = LiberacionIngles.objects.get_or_create(periodo=str(header_row[1].value), alumno=alumno)
                if created:
                    results['created'] += 1
            except Alumno.DoesNotExist as ex:
                results['errors'].append({'type': str(type(ex)), 'message': f'No se encontro un alumno con no. de control {data["no_control"]}', 'row_index': row[0].row})
            except Exception as ex:
                results['errors'].append({'type': str(type(ex)), 'message': str(ex), 'row_index': row[0].row})
        return Response(status=200, data=results)

### CORTE
@api_view(['POST',])
@permission_classes([IsAuthenticated&IsAdminUser])
def corte(request):
    periodo = getPeriodoActual()
    # Verificar si no existen registros que pertenezcan a un corte para el periodo actual
    if not Ingreso.objects.contiene_corte(periodo) and not Egreso.objects.contiene_corte(periodo) and not Titulacion.objects.contiene_corte(periodo) and not LiberacionIngles.objects.contiene_corte(periodo):
        # Realizar corte
        ingresos = Ingreso.objects.realizar_corte(periodo)
        egresos = Egreso.objects.realizar_corte(periodo)
        titulaciones = Titulacion.objects.realizar_corte(periodo)
        liberaciones = LiberacionIngles.objects.realizar_corte(periodo)
        return Response(status=200, data={'periodo': periodo, 'updated': {'ingresos': ingresos, 'egresos': egresos, 'titulaciones': titulaciones, 'liberaciones-ingles': liberaciones}})
    else:
        return Response(status=400, data={'periodo': periodo ,'message': f'No se puede realizar un corte ya que existen registros que pertenecen a un corte para el periodo {periodo}.'})

# Esta función convierte una fila de Excel en un diccionario basado en los encabezados
def row_to_dict(header_row, data_row):
    # Limpia las filas eliminando celdas vacías
    clean_header = clean_row(header_row)  # Limpia la fila de encabezados
    clean_data = clean_row(data_row)      # Limpia la fila de datos
    
    # Lista de campos esperados en el archivo Excel
    keywords = ['curp', 'no_control', 'paterno', 'materno', 'nombre', 'carrera']
    
    # Inicializa el diccionario con una lista vacía para los periodos
    row_dict = {'periodos': []}
    
    # Procesa cada celda en la fila de datos
    for cell in clean_data:
        # Obtiene el índice de la columna (restando 1 porque Excel empieza en 1)
        index = cell.column - 1
        # Convierte el encabezado a minúsculas
        header = str(clean_header[index].value).lower()
        # Obtiene el valor de la celda, si es None retorna None
        value = str(cell.value) if cell.value else None
        
        # Si el encabezado es uno de los campos esperados
        if header in keywords:
            row_dict[header] = value
        # Si el encabezado es un periodo (formato: YYYYS donde Y=año, S=semestre)
        elif re.match(r'^[12][0-9]{3}[13]$', header):
            # Si hay un valor, agrega una tupla (periodo, tipo) a la lista de periodos
            if value: row_dict['periodos'].append((header, value[0:2]))
        else:
            # Si el encabezado no es reconocido, lanza una excepción
            raise Exception(f'Campo "{header}" no es reconocido')
    
    # Ordena los periodos cronológicamente
    row_dict['periodos'].sort(key=lambda x: x[0])
    return row_dict

# Esta función limpia una fila de Excel eliminando las celdas vacías
def clean_row(row):
    clean = []
    # Recorre cada celda en la fila
    for cell in row:
        # Si la celda tiene un valor, la agrega a la lista limpia
        if cell.value is not None:
            clean.append(cell)
    return clean